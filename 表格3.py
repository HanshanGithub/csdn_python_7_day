import openpyxl as op
file = "D:/AmStorage/My_File/SCDN_test/file/grade.xlsx"
# 加载工作薄文件
wb = op.load_workbook(file)
# 获取工作表信息
wb_list = wb.sheetnames
# 获取第一个工作表信息
sht0 = wb[wb_list[0]]
# 得到工作表的区间
sht0.dimensions
# 指定单元格的值
sht0['A2'].value
sht0.cell(2,2).value
# 输出
# 单元格区间的引用
area1 = sht0['A1:C3']
# 输出单元格区间的各值，先遍历行再遍历列
for cell_rows in area1:
    for cell_cols in cell_rows:
        print(cell_cols.value)

wb.close()
